import pandas as pd
import openpyxl
import os

# CONFIGURATION
CSV_PATH = 'Species_params_GEMINI_populated.csv'
MASTER_XLSM = 'Weighted_species_params_master.xlsm'
CHILD_XLSX = 'Weighted_species_params_child.xlsx'
TEMP_MASTER = 'Weighted_species_params_master_temp.xlsm'

# Mapping XLSX column headers to CSV scientific names
# Note: Using the column names exactly as they appear in the XLSM
SPECIES_MAP = {
    'cynodon_dactylon': 'Cynodon dactylon',
    'cenchrus_mezianum': 'cenchrus mezianum or pennisetum mezianum',
    'themeda_triandra': 'Themeda triandra',
    'indigofera_volkensii': 'Indigofera volkensii',
    'vechellia_drepanolobium': 'Vachellia drepanolobium',
    'tephrosia_pumilla': 'tephrosia pumila'
}

def sync_data():
    print(f"Loading source data from {CSV_PATH}...")
    df_csv = pd.read_csv(CSV_PATH)
    
    # Create a lookup: (scientific_name, code) -> full row data
    lookup = {}
    for _, row in df_csv.iterrows():
        lookup[(row['scientific_name'], row['code'])] = row

    print(f"Loading master workbook: {MASTER_XLSM}...")
    try:
        wb = openpyxl.load_workbook(MASTER_XLSM, keep_vba=True)
    except Exception as e:
        print(f"CRITICAL ERROR: Could not load {MASTER_XLSM}. Ensure the file exists and is not corrupt. {e}")
        return

    ws = wb.active # Usually Sheet1
    
    # Map headers to column indices (1-based)
    header_row = [cell.value for cell in ws[1]]
    headers_idx = {val: idx for idx, val in enumerate(header_row, start=1)}
    
    # Verify all mapped species exist in the spreadsheet
    active_col_map = {}
    for xlsx_col, csv_name in SPECIES_MAP.items():
        if xlsx_col in headers_idx:
            active_col_map[xlsx_col] = (headers_idx[xlsx_col], csv_name)
        else:
            print(f"  Warning: Column '{xlsx_col}' not found in XLSM. Skipping.")

    print("Populating data with MAX/MIN logic...")
    updated_count = 0
    code_col_idx = headers_idx.get('code', 1)
    
    # Iterate through rows starting from row 2
    for row_idx in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row_idx, column=code_col_idx).value
        if not code_val:
            continue
        
        clean_code = str(code_val).strip().rstrip(',')
        
        for xlsx_col, (col_idx, csv_name) in active_col_map.items():
            row_data = lookup.get((csv_name, clean_code))
            
            # Handle known synonyms (e.g. AMX25/AMAXB)
            if row_data is None:
                if clean_code == 'AMX25':
                    row_data = lookup.get((csv_name, 'AMAXB'))
                elif clean_code == 'AMAXB':
                    row_data = lookup.get((csv_name, 'AMX25'))
                elif clean_code == 'VCMAX25':
                    row_data = lookup.get((csv_name, 'VCMAX'))
                elif clean_code == 'VCMAX':
                    row_data = lookup.get((csv_name, 'VCMAX25'))
                elif clean_code == 'GDD_EMERG':
                    row_data = lookup.get((csv_name, 'GDD_EMERGENCE'))
                elif clean_code == 'GDD_EMERGENCE':
                    row_data = lookup.get((csv_name, 'GDD_EMERG'))
            
            if row_data is not None:
                matched_code = str(row_data['code']).upper()
                
                # Apply conditional logic
                if 'MAX' in matched_code:
                    val = row_data['maximum']
                elif 'MIN' in matched_code:
                    val = row_data['minimum']
                else:
                    val = row_data['midpoint']
                
                if pd.notna(val):
                    ws.cell(row=row_idx, column=col_idx).value = val
                    updated_count += 1

    # Save Master
    print(f"Saving changes to master...")
    try:
        wb.save(TEMP_MASTER)
        if os.path.exists(MASTER_XLSM):
            os.remove(MASTER_XLSM)
        os.rename(TEMP_MASTER, MASTER_XLSM)
        print(f"SUCCESS: {MASTER_XLSM} updated with {updated_count} parameters.")
    except Exception as e:
        print(f"ERROR: Could not finalize master update. {e}")
        print(f"Temporary file saved as: {TEMP_MASTER}")
        return

    # Create Child
    print(f"Synchronizing child file: {CHILD_XLSX}...")
    try:
        # Load master (without keep_vba) to save a clean xlsx
        wb_child = openpyxl.load_workbook(MASTER_XLSM)
        wb_child.save(CHILD_XLSX)
        print(f"SUCCESS: {CHILD_XLSX} created/updated.")
    except Exception as e:
        print(f"ERROR: Could not update child file. {e}")

if __name__ == "__main__":
    sync_data()
