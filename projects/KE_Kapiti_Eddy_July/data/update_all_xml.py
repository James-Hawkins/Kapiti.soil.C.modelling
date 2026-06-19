import openpyxl
import re
import os

xlsm_path = 'Weighted_species_params_master.xlsm'
species_xml_path = '../KE_Kapiti_speciesparameters_GLOBAL.xml'
events_xml_path = '../KE_Kapiti_events_eddy.xml'

def format_value(v):
    try:
        f_val = float(v)
        if f_val.is_integer(): return str(int(f_val))
        return str(round(f_val, 6))
    except:
        return str(v)

def main():
    if not os.path.exists(xlsm_path):
        print(f"Error: {xlsm_path} not found.")
        return

    print("Loading XLSM...")
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    code_idx = headers.index('code')
    
    try:
        grass_idx = headers.index('HYBRID.GRASS')
        forb_idx = headers.index('HYBRID.FORB')
        shrub_idx = headers.index('vechellia_drepanolobium')
    except ValueError as e:
        print(f"Header missing: {e}")
        return

    grass_species = {}
    forb_species = {}
    grass_events = {}
    forb_events = {}
    shrub_events = {}

    event_keys = ['initialbiomass', 'heightmax', 'rootingdepth', 'dbh', 'treenumber']
    ignore_keys = ['fractionalcover', 'weight.all', 'weight.ft', 'biomass.yield.Mg.ha.yr']

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[code_idx]
        if not code: continue
        code = str(code).strip()
        
        g_val = row[grass_idx]
        if g_val is not None and str(g_val).strip() != '' and str(g_val).lower() not in ('nan', 'na'):
            if code in event_keys:
                grass_events[code] = format_value(g_val)
            elif code not in ignore_keys and not code.startswith('weight'):
                grass_species[code] = format_value(g_val)
                
        f_val = row[forb_idx]
        if f_val is not None and str(f_val).strip() != '' and str(f_val).lower() not in ('nan', 'na'):
            if code in event_keys:
                forb_events[code] = format_value(f_val)
            elif code not in ignore_keys and not code.startswith('weight'):
                forb_species[code] = format_value(f_val)

        s_val = row[shrub_idx]
        if s_val is not None and str(s_val).strip() != '' and str(s_val).lower() not in ('nan', 'na'):
            if code in event_keys:
                shrub_events[code] = format_value(s_val)

    print(f"Extracted Grass: {len(grass_species)} species params, {len(grass_events)} event params.")
    print(f"Extracted Forb: {len(forb_species)} species params, {len(forb_events)} event params.")
    print(f"Extracted Shrub: {len(shrub_events)} event params.")

    with open(species_xml_path, 'r', encoding='utf-8') as f:
        species_xml = f.read()

    def update_species_block(content, mnemonic, params):
        pattern = r'(<species[^>]*mnemonic="' + re.escape(mnemonic) + r'"[^>]*>)(.*?)(</species>)'
        match = re.search(pattern, content, re.DOTALL)
        if not match:
            print(f"Warning: Block for {mnemonic} not found in Species XML.")
            return content, 0
        
        start_tag = match.group(1)
        inner = match.group(2)
        end_tag = match.group(3)
        
        updates = 0
        for k, v in params.items():
            par_pattern = r'(<par\s+name="' + re.escape(k) + r'"\s+value=")[^"]*(".*?>)'
            if re.search(par_pattern, inner):
                inner = re.sub(par_pattern, r'\g<1>' + v + r'\g<2>', inner)
            else:
                inner += f'        <par name="{k}" value="{v}"/>\n'
            updates += 1
        
        new_block = start_tag + inner + end_tag
        return content[:match.start()] + new_block + content[match.end():], updates

    species_xml, gs_up = update_species_block(species_xml, 'GRASS.HYBRID.1', grass_species)
    species_xml, fs_up = update_species_block(species_xml, 'FORB.HYBRID.1', forb_species)

    with open(species_xml_path, 'w', encoding='utf-8') as f:
        f.write(species_xml)
    print(f"Species XML updated: {gs_up} grass values, {fs_up} forb values.")

    with open(events_xml_path, 'r', encoding='utf-8') as f:
        events_xml = f.read()

    def update_events_block(content, plant_type, params):
        pattern = r'(<plant\s+type="' + re.escape(plant_type) + r'"\s*>\s*<[a-zA-Z0-9_]+)([^>]*)(/>|</[a-zA-Z0-9_]+>)'
        
        def repl(m):
            tag_start = m.group(1)
            attrs = m.group(2)
            tag_end = m.group(3)
            
            for k, v in params.items():
                attr_pattern = r'(' + re.escape(k) + r'\s*=\s*")[^"]*(")'
                if re.search(attr_pattern, attrs):
                    attrs = re.sub(attr_pattern, r'\g<1>' + v + r'\g<2>', attrs)
                else:
                    attrs += f' {k}="{v}"'
            return tag_start + attrs + tag_end

        return re.sub(pattern, repl, content, flags=re.DOTALL)

    events_xml = update_events_block(events_xml, 'GRASS.HYBRID.1', grass_events)
    events_xml = update_events_block(events_xml, 'FORB.HYBRID.1', forb_events)
    events_xml = update_events_block(events_xml, 'HYBRID.SHRUB', shrub_events)

    with open(events_xml_path, 'w', encoding='utf-8') as f:
        f.write(events_xml)
    print("Events XML updated.")

if __name__ == '__main__':
    main()
