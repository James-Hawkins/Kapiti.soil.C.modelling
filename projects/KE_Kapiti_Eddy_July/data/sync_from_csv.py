import openpyxl
import csv
import os

source_csv = 'Species_params_GEMINI_populated.csv'
target_xlsm = 'Weighted_species_params_master.xlsm'

print(f"Syncing {source_csv} to {target_xlsm}...")

def normalize_species(name):
    if not name: return ""
    n = str(name).lower().strip()
    if "themeda" in n: return "themeda_triandra"
    if "cynodon" in n: return "cynodon_dactylon"
    if "cenchrus" in n: return "cenchrus_mezianum"
    if "indigofera" in n: return "indigofera_volkensii"
    if "tephrosia" in n: return "tephrosia"
    if "vachellia" in n: return "vachellia_drepanolobium"
    return n.replace(" ", "_")

def normalize_code(code):
    if not code: return ""
    return str(code).strip().strip('"').strip("'").strip(',').strip()

# Read source CSV
extracted_data = []
with open(source_csv, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        code = row.get('code')
        if not code: code = row.get('Code')
        spec = row.get('scientific_name')
        if not spec: spec = row.get(list(row.keys())[2])
        
        mid = row.get('midpoint')
        mx = row.get('maximum')
        mn = row.get('minimum')
        
        if mid is not None and str(mid).strip() != '' and str(mid).lower() != 'nan' and str(mid) != 'None':
            extracted_data.append({
                'code': normalize_code(code),
                'species': normalize_species(spec),
                'min': mn,
                'max': mx,
                'mid': mid
            })

print(f"Extracted {len(extracted_data)} populated parameters from CSV.")

try:
    wb_tgt = openpyxl.load_workbook(target_xlsm, keep_vba=True)
    ws_tgt = wb_tgt.active
except Exception as e:
    print(f"Error loading {target_xlsm}: {e}")
    exit(1)

headers_tgt = [cell.value for cell in ws_tgt[1]]
h_map = {val: i+1 for i, val in enumerate(headers_tgt) if val}

for h in list(h_map.keys()):
    hl = str(h).lower()
    if "tephrosia" in hl: h_map["tephrosia"] = h_map[h]
    if "vachellia" in hl: h_map["vachellia_drepanolobium"] = h_map[h]
    if "indigofera" in hl: h_map["indigofera_volkensii"] = h_map[h]

row_map = {}
for r in range(2, ws_tgt.max_row + 1):
    c = ws_tgt.cell(row=r, column=1).value
    if c: row_map[str(c).strip()] = r

updates = 0
for data in extracted_data:
    sp = data['species']
    c = data['code']
    if not sp or not c: continue

    v = ""
    if c.endswith("MAX") or c.endswith("MAX25") or c.endswith("VCMAX25"):
        v = data['max']
    elif c.endswith("MIN"):
        v = data['min']
    else:
        v = data['mid']

    if v is None or str(v).strip() == "" or str(v).strip().lower() == "na" or str(v) == 'None':
        continue

    v = str(v).strip()

    if sp in h_map:
        tgt_r = row_map.get(c)
        if not tgt_r:
            for tc, tr in row_map.items():
                if tc.startswith(c) or c.startswith(tc):
                    tgt_r = tr
                    break
        
        if tgt_r:
            try:
                num_v = float(v)
            except:
                num_v = v
            ws_tgt.cell(row=tgt_r, column=h_map[sp], value=num_v)
            updates += 1

wb_tgt.save(target_xlsm)
print(f"Applied {updates} updates to {target_xlsm}")
