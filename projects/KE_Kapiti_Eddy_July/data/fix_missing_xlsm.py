import openpyxl

file_path = 'Species_params_GEMINI.xlsm'
wb = openpyxl.load_workbook(file_path, keep_vba=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
mid_col = headers.index('midpoint') + 1

# Cynodon
cynodon_data = {
    89: 0.0,
    90: 0.025,
    91: 0.010,
    92: 0.007,
    93: 0.0115,
    94: 0.0055,
    95: 1.25
}

# Cenchrus
cenchrus_data = {
    136: 0.0,
    137: 0.020,
    138: 0.0105,
    139: 0.007,
    140: 0.010,
    141: 0.00725,
    142: 1.75
}

updates = 0
for r, val in cynodon_data.items():
    ws.cell(row=r, column=mid_col, value=val)
    updates += 1

for r, val in cenchrus_data.items():
    ws.cell(row=r, column=mid_col, value=val)
    updates += 1

wb.save(file_path)
print(f"Directly updated {updates} cells in XLSM for Bermuda and African Foxtail.")
