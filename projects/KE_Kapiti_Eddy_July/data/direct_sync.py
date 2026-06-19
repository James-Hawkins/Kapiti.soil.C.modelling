import openpyxl
import os

source_xlsm = 'Species_params_GEMINI.xlsm'
target_xlsm = 'Weighted_species_params_master.xlsm'

def normalize_species(name):
    if not name: return ""
    n = str(name).lower().strip()
    if "themeda triandra" in n: return "themeda_triandra"
    if "cynodon" in n: return "cynodon_dactylon"
    if "cenchrus" in n: return "cenchrus_mezianum"
    if "indigofera" in n: return "indigofera_volkensii"
    if "tephrosia" in n: return "tephrosia"
    if "vachellia" in n: return "vachellia_drepanolobium"
    return n.replace(" ", "_")

def normalize_code(code):
    if not code: return ""
    return str(code).strip().strip('"').strip("'").strip(',').strip()

print("Extracting data from source XLSM...")
wb_src = openpyxl.load_workbook(source_xlsm, data_only=True)
ws_src = wb_src.active

headers_src = [cell.value for cell in ws_src[1]]
c_code = headers_src.index('code') + 1 if 'code' in headers_src else headers_src.index('Code') + 1
c_spec = headers_src.index('scientific_name') + 1 if 'scientific_name' in headers_src else 3
c_min = headers_src.index('minimum') + 1
c_max = headers_src.index('maximum') + 1
c_mid = headers_src.index('midpoint') + 1

extracted_data = []

# Because data_only=True evaluates formulas, codes like =J42 might evaluate to None if Excel hasn't calculated them.
# So we need to handle that. If code is None, maybe we can figure it out, but wait, earlier we found they DO evaluate correctly or they were formulas.
# Wait, for Vachellia, earlier we found the code was '=J136' with data_only=False, but it didn't evaluate with data_only=True.
# So we must use data_only=False to get the code if it's a formula? Or just read the string value.
wb_src_vba = openpyxl.load_workbook(source_xlsm, keep_vba=True)
ws_src_vba = wb_src_vba.active

for r in range(2, ws_src.max_row + 1):
    # Try to get code from data_only
    code = ws_src.cell(row=r, column=c_code).value
    # If code is None or looks like a formula, try to trace it or just hardcode the known block sizes
    # Actually, if we just use the original populated CSV we merged earlier...
    # Let's just use the logic from earlier:
    
    # Let's just read the code from the corresponding template row
    # If it's a formula like =J42, we know the code.
    code_vba = str(ws_src_vba.cell(row=r, column=c_code).value)
    if code_vba.startswith("="):
        # Extract the row number
        ref_row = int(''.join(filter(str.isdigit, code_vba)))
        code = str(ws_src_vba.cell(row=ref_row, column=c_code).value)
    elif code_vba != 'None':
        code = code_vba

    spec = ws_src.cell(row=r, column=c_spec).value
    # Sometimes species is also a formula
    spec_vba = str(ws_src_vba.cell(row=r, column=c_spec).value)
    if spec_vba.startswith("="):
        ref_row = int(''.join(filter(str.isdigit, spec_vba)))
        spec = str(ws_src_vba.cell(row=ref_row, column=c_spec).value)
    elif spec_vba != 'None':
        spec = spec_vba

    mid = ws_src.cell(row=r, column=c_mid).value
    mx = ws_src.cell(row=r, column=c_max).value
    mn = ws_src.cell(row=r, column=c_min).value
    
    if mid is not None and str(mid).strip() != '' and str(mid).lower() != 'nan' and str(mid) != 'None':
        extracted_data.append({
            'code': str(code).strip() if code else "",
            'species': str(spec).strip() if spec else "",
            'min': mn,
            'max': mx,
            'mid': mid
        })

print(f"Extracted {len(extracted_data)} parameters.")

print("Updating Target XLSM...")
wb_tgt = openpyxl.load_workbook(target_xlsm, keep_vba=True)
ws_tgt = wb_tgt.active

headers_tgt = [cell.value for cell in ws_tgt[1]]
h_map = {val: i+1 for i, val in enumerate(headers_tgt) if val}

for h in list(h_map.keys()):
    hl = str(h).lower()
    if "tephrosia" in hl: h_map["tephrosia"] = h_map[h]
    if "vachellia" in hl: h_map["vachellia_drepanolobium"] = h_map[h]
    if "indigofera" in hl: h_map["indigofera_volkensii"] = h_map[h]

row_map = {}
for r in range(2, ws_tgt.max_row + 1):
    c = ws_tgt.cell(row=r, column=1).value
    if c: row_map[str(c).strip()] = r

updates = 0
for data in extracted_data:
    sp = normalize_species(data['species'])
    c = str(data['code']).strip()
    if not sp or not c: continue

    v = ""
    if c.endswith("MAX") or c.endswith("MAX25") or c.endswith("VCMAX25"):
        v = data['max']
    elif c.endswith("MIN"):
        v = data['min']
    else:
        v = data['mid']

    if v is None or str(v).strip() == "" or str(v).strip().lower() == "na" or str(v) == 'None':
        continue

    v = str(v).strip()

    if sp in h_map:
        tgt_r = row_map.get(c)
        if not tgt_r:
            for tc, tr in row_map.items():
                if tc.startswith(c) or c.startswith(tc):
                    tgt_r = tr
                    break
        
        if tgt_r:
            try:
                num_v = float(v)
            except:
                num_v = v
            ws_tgt.cell(row=tgt_r, column=h_map[sp], value=num_v)
            updates += 1

wb_tgt.save(target_xlsm)
print(f"Applied {updates} updates to {target_xlsm}")
