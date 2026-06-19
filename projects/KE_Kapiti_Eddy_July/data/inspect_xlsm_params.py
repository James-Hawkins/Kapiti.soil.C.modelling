import openpyxl
import os

file_path = 'Species_params_GEMINI.xlsm'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    code_idx = None
    if 'code' in headers:
        code_idx = headers.index('code')
    elif 'Code' in headers:
        code_idx = headers.index('Code')

    if code_idx is not None:
        target_params = [
            'NFIX_RATE', 'NCFOLOPT', 'NC_FINEROOTS_MAX', 'NC_FINEROOTS_MIN', 
            'NC_STRUCTURAL_TISSUE_MAX', 'NC_STRUCTURAL_TISSUE_MIN', 'QRF'
        ]
        
        species_idx = headers.index('scientific_name') if 'scientific_name' in headers else 2
        search_idx = headers.index('search.text') if 'search.text' in headers else 7
        
        print(f"Target parameters to extract: {target_params}")
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            code = str(row[code_idx]).strip() if row[code_idx] else ""
            if code in target_params:
                species = row[species_idx]
                search_text = row[search_idx]
                print(f"Row {r_idx} | Species: {species} | Code: {code}")
                print(f"Search Text: {search_text}\n")
    else:
        print("Could not find 'code' column.")

except Exception as e:
    print(f"Error: {e}")
